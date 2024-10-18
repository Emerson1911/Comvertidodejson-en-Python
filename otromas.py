import pandas as pd
import openpyxl
import re
import json

def normalize_text(text):
    """
    Normaliza el texto y extrae solo el segundo comentario si existe
    """
    if not text:
        return ""
    
    # Dividir por saltos de línea o punto y coma para separar comentarios
    parts = re.split(r'[\n;]+', str(text))
    
    # Tomar el segundo comentario si existe, si no, usar el texto completo
    text_to_process = parts[1].strip() if len(parts) > 1 else parts[0].strip()
    
    # Normalizar el texto
    text_to_process = text_to_process.upper()
    replacements = {
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'Ü': 'U', 'Ñ': 'N'
    }
    for k, v in replacements.items():
        text_to_process = text_to_process.replace(k, v)
    return text_to_process

def extract_number(text):
    """
    Extrae el primer número encontrado en el texto
    """
    if not text:
        return None
    numbers = re.findall(r'\d+\.?\d*', str(text))
    return float(numbers[0]) if numbers else None

def find_route_in_comment(comment, cities):
    """
    Busca ciudades de origen y destino en el comentario
    """
    normalized_comment = normalize_text(comment)
    found_cities = []
    
    for city in cities:
        if normalize_text(city) in normalized_comment:
            found_cities.append(city)
    
    return found_cities

def get_second_comment(comment_text):
    """
    Extrae el segundo comentario del texto si existe
    """
    parts = re.split(r'[\n;]+', str(comment_text))
    return parts[1].strip() if len(parts) > 1 else parts[0].strip()

def analyze_routes(file_path, cities, sheet_name=None):
    """
    Analiza el archivo Excel para encontrar rutas entre ciudades
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active if sheet_name is None else wb[sheet_name]
    
    routes = []
    
    # Buscar en todas las celdas
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            
            # Si la celda tiene un comentario y un valor numérico
            if cell.comment and extract_number(cell.value) is not None:
                comment_text = cell.comment.text.strip()
                kilometers = extract_number(cell.value)
                
                # Extraer solo el segundo comentario si existe
                processed_comment = get_second_comment(comment_text)
                
                # Encontrar ciudades mencionadas en el comentario
                cities_in_comment = find_route_in_comment(processed_comment, cities)
                
                # Buscar la ciudad de origen (generalmente está en la columna)
                origin_city = None
                for search_row in range(row-5, row):  # Buscar hasta 5 celdas arriba
                    if search_row > 0:
                        header_cell = sheet.cell(row=search_row, column=col)
                        if header_cell.value:
                            cities_in_header = find_route_in_comment(str(header_cell.value), cities)
                            if cities_in_header:
                                origin_city = cities_in_header[0]
                                break
                
                # Si encontramos origen y al menos una ciudad en el comentario
                if origin_city and cities_in_comment:
                    # Filtrar la ciudad de origen del comentario si está presente
                    destination_cities = [city for city in cities_in_comment if city != origin_city]
                    
                    if destination_cities:  # Si hay ciudades destino
                        for dest_city in destination_cities:
                            routes.append({
                                "origen": origin_city,
                                "destino": dest_city,
                                "kilometraje": kilometers,
                                "comentario": processed_comment,  # Usar solo el segundo comentario
                                "columna": sheet.cell(row=1, column=col).value or f"Columna {col}"
                            })
    
    # Ordenar las rutas por origen y destino
    routes.sort(key=lambda x: (x["origen"], x["destino"]))
    
    # Convertir a JSON
    json_data = json.dumps(routes, indent=4, ensure_ascii=False)
    
    # Guardar el JSON
    output_file = "rutas_detalladas.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(json_data)
    
    return json_data

# Lista de ciudades
ciudades = [
    "AGUADULCE",
    "AHUACHAPAN",
    "ALAJUELA",
    "ALTA VERAPAZ",
    "AMATITLAN",
    "ANGUIATU",
    "ANTIGUA GUATEMALA",
    "ATLANTIDA",
    "BAJA VERAPAZ",
    "BELMOPAN",
    "BOACO",
    "BOCA DEL TORO",
    "BUENOS AIRES",
    "CABAÑAS",
    "CARAZO",
    "CARTAGO",
    "CATACAMAS",
    "CAÑAS",
    "CHALATENANGO",
    "CHAMPERICO",
    "CHANGUINOLA",
    "CHIMALTENANGO",
    "CHINANDEGA",
    "CHIQUIMULA",
    "CHIRIQUI",
    "CHITRE",
    "CHOLOMA",
    "CHOLUTECA",
    "CHONTALES",
    "CIUDAD DE GUATEMALA",
    "CIUDAD DE PANAMA",
    "CIUDAD HIDALGO",
    "CIUDAD QUEZADA",
    "CIUDAD TECUN UMAN",
    "COBAN",
    "COCLE",
    "COJUTEPEQUE",
    "COLON",
    "COLON HONDURAS",
    "COMAYAGUA",
    "COPAN",
    "CORTES",
    "CUILAPA",
    "CUSCATLAN",
    "DANLI",
    "DAVID",
    "DIRIAMBA",
    "EL PARAISO",
    "EL PROGRESO",
    "EL PROGRESO YORO",
    "EL ROBLE",
    "ESCUINTLA",
    "ESPARTA",
    "ESPARZA",
    "ESTELI",
    "FLORES",
    "FLORES DE LEAN",
    "FRANCISCO MORAZAN",
    "FRONTERA AGUA CALIENTE",
    "FRONTERA CORINTO",
    "FRONTERA EL AMATILLO",
    "FRONTERA EL CEIBO",
    "FRONTERA EL ESPINO",
    "FRONTERA EL FLORIDO",
    "FRONTERA EL POY",
    "FRONTERA GUASAULE",
    "FRONTERA LA ERMITA",
    "FRONTERA LA FRATERNIDAD",
    "FRONTERA LA HACHADURA",
    "FRONTERA LAS CHINAMAS",
    "FRONTERA LAS MANOS",
    "FRONTERA PASO CANOAS",
    "FRONTERA PEDRO DE ALVARADO",
    "FRONTERA PEÑAS BLANCAS",
    "FRONTERA SAN CRISTOBAL",
    "FRONTERA VALLE NUEVO",
    "GRACIAS",
    "GRANADA",
    "GUANACASTE",
    "GUAPILES",
    "GUASTATOYA",
    "HEREDIA",
    "HERRERA",
    "HUEHUETENANGO",
    "INTIBUCA",
    "IXCAN",
    "IZABAL",
    "JALAPA",
    "JINOTEGA",
    "JINOTEPE",
    "JUIGALPA",
    "JUTIAPA",
    "JUTICALPA",
    "LA CEIBA",
    "LA CHORRERA",
    "LA CRUZ",
    "LA ESPERANZA",
    "LA LIBERTAD AREA PUERTO",
    "LA LIBERTAD AREA SITIO DEL NIÑO",
    "LA PAZ EL SALVADOR",
    "LA PAZ HONDURAS",
    "LA UNION",
    "LAS TABLAS",
    "LEMPIRA",
    "LEON",
    "LIBERIA",
    "LIMON",
    "LOS SANTOS",
    "MADRIZ",
    "MANAGUA",
    "MASAYA",
    "MATAGALPA",
    "MAZATENANGO",
    "MELCHOR DE MENCOS",
    "METAPAN",
    "MIXCO",
    "MORALES",
    "MORAZAN",
    "NACAOME",
    "NANDAIME",
    "NICOYA",
    "NUEVA SEGOVIA",
    "OCOTAL",
    "OCOTEPEQUE",
    "OLANCHITO",
    "OLANCHO",
    "PALIN",
    "PALMEROLA",
    "PENOMOME",
    "PETEN",
    "PUERTO BARRIOS",
    "PUERTO CALDERA",
    "PUERTO CASTILLA",
    "PUERTO CORINTO",
    "PUERTO CORTES",
    "PUERTO DE ACAJUTLA",
    "PUERTO DE LA UNION",
    "PUERTO LIMON",
    "PUERTO QUETZAL",
    "PUERTO SANDINO",
    "PUERTO SANTO TOMAS DE CASTILLA",
    "PUNTARENAS",
    "QUETZALTENANGO",
    "QUICHE",
    "RETALHULEU",
    "RIO SAN JUAN",
    "RIVAS",
    "SABA",
    "SALAMA",
    "SAN CARLOS",
    "SAN FRANCISCO GOTERA",
    "SAN ISIDRO DE EL GENERAL",
    "SAN JOSE COSTA RICA",
    "SAN LORENZO",
    "SAN MARCOS",
    "SAN MIGUEL",
    "SAN PEDRO SOLOMA",
    "SAN PEDRO SULA",
    "SAN RAFAEL CEDROS",
    "SAN SALVADOR",
    "SAN VICENTE",
    "SANTA ANA",
    "SANTA BARBARA",
    "SANTA CRUZ",
    "SANTA CRUZ DEL QUICHE",
    "SANTA ROSA",
    "SANTA ROSA DE COPAN",
    "SANTIAGO",
    "SENSUNTEPEQUE",
    "SIGUATEPEQUE",
    "SIQUIRES",
    "SOLOLA",
    "SOMOTO",
    "SONSONATE",
    "TEGUCIGALPA",
    "TELA",
    "TOCOA",
    "TOTONICAPAN",
    "TRUJILLO",
    "USULUTAN",
    "VALLE",
    "VERAGUAS",
    "VILLA NUEVA GUATEMALA",
    "VILLANUEVA HONDURAS",
    "YORO",
    "YUSCARAN",
    "ZACAPA",
    "ZACATEPEQUEZ"
]

# Uso del script
try:
    file_path = "C:\\Users\\emers\\Desktop\\Matriz de ruta.xlsx"  # Ajusta esta ruta a tu archivo
    json_result = analyze_routes(file_path, ciudades)
    print("¡Conversión exitosa!")
    print("\nMuestra del JSON generado:")
    print(json_result[:500] + "...")
    
except Exception as e:
    print(f"Error: {str(e)}")